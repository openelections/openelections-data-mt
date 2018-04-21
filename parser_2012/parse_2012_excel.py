from openpyxl import load_workbook
import unicodecsv

outfile = '20121106__mt__general__precinct.csv'

# normalize precinct names
# get rid of #, convert abbreviations to PRECINCT
def cleanup_precinct(orig):
  s = orig 
  s = s.replace("PREC ", "PRECINCT ")
  s = s.replace("PRECNCT", "PRECINCT")
  s = s.replace("PCT", "PRECINCT")
  s = s.replace("#", "")
  #if s != orig:
  #  print(f'Cleanup precinct: orig {orig} to {s}')
  return(s)

def parse_all():
  headers = ['county', 'precinct', 'office', 'district', 'party', 'candidate', 'votes']
  #counties = get_counties()

  with open('countylist.txt') as f:
     counties = f.read().splitlines()
  #print("COUNTIES: {counties}")

  with open(outfile, 'wb') as csvfile:
      w = unicodecsv.writer(csvfile, encoding='utf-8')
      w.writerow(headers)

      for c in counties:
        county = c.split(" Results")[0]
        data_file = "data/" + c
        print(f'PROCESSING county {county} file {data_file}')
        parse_county_excel(w, data_file, county)


def parse_county_excel(w, data_file, county):

  book = load_workbook(data_file)
  sheets = book.get_sheet_names()
  total_sheets = len(sheets)

  hdr_row = 7
  for sheet in sheets:
    sh = book.get_sheet_by_name(sheet)

    header = [sh.cell(row=7, column=x).value for x in range(1, sh.max_column+1)]

    district = ''
    party = ''

    office = header[0]
    if office.find(" OF DISTRICT COURT") < 0 and office.find(" DISTRICT") > 0:
      #print(f'FOUND DISTRICT: {office}')
      zz = office.split(" DISTRICT ")
      if(len(zz) == 2):
        office = zz[0].strip().strip(",")
        district = zz[1].strip()
        #print(f'   => {office} ==> {district}')

    office = office.replace("\n", " ")
    if office.find("Shall") >= 0:
      #print(f'Shall {office}')
      d = office.find(" of District ")
      if d >= 0:
        tmp = office.split(" of District ")[1].split(",")[0]
        #print(f'   ===>  {tmp}')
        district = tmp

    candidates = header[2:]

    for r in range(8, sh.max_row + 1):
      row = [sh.cell(row=r, column=x).value for x in range(1, sh.max_column+1)]
      precinct = cleanup_precinct(row[1])
      
      if precinct != 'TOTALS':
        cand_votes = zip(candidates, [x for x in row[2:]])
        for cand in cand_votes:
          name = cand[0]
          vote = int(cand[1])
          nl = name.find("\n")
          s = name
          if nl > 0:
            name = s[:nl]
            party = s[(nl+1): ]
          else:
            party = '' 
           
          w.writerow([county, precinct, office, district, party, name, vote])


def main():
  parse_all()

if __name__ == "__main__":
    main()
